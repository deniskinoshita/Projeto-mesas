"""
/api/pdf (gerar_pdf) e /api/xlsx (gerar_xlsx) — geração de documentos para o
assessor a partir dos mesmos dados de composição/desvio.

Passa pelas rotas Flask (test_client), não pelas funções isoladas: são elas
que normalizam o payload (JSON solto do front) antes de chamar gerar_pdf/
gerar_xlsx, e é esse contrato que quebra em produção quando alguém mexe na
rota sem mexer na função (ou vice-versa).

Para o PDF: confere que gera sem exceção, com content-type e magic bytes
corretos (não dá pra fazer assert fino de layout render de reportlab sem um
parser de PDF dedicado — fora do escopo aqui).

Para o XLSX: reabre o arquivo gerado com openpyxl e confere que os VALORES
das células batem com o que foi enviado (não só "não quebrou") — patrimônio,
caixa, e a linha de alocação por classe (atual/modelo/desvio/valor do gap).
"""
import io

import pytest
from openpyxl import load_workbook


PAYLOAD_BASE = {
    "nome": "Cliente Teste",
    "perfil": "moderada",
    "conta": "123456",
    "assessor": "Denis Kinoshita",
    "modelo_nome": "Levante Asset",
    "data_ref": "15/07/2026",
    "patrimonio": 100_000.0,
    "caixa": 5.0,
    "desvios": [
        {"cat": "acoes", "label": "Ações", "real": 20.0, "alvo": 10.0, "desvio": 10.0},
        {"cat": "pos_fixado", "label": "Pós Fixado", "real": 30.0, "alvo": 44.0, "desvio": -14.0},
    ],
    "rent": {},
    "macro": {},
    "recomendacoes": [],
    "alertas": [],
    "ref_contexto": "",
    "acoes": [],
}


def test_api_pdf_gera_documento_sem_excecao(client):
    resp = client.post("/api/pdf", json=PAYLOAD_BASE)
    assert resp.status_code == 200
    assert resp.mimetype == "application/pdf"
    assert resp.data[:4] == b"%PDF"  # magic bytes de um PDF válido
    assert len(resp.data) > 1000  # documento não-trivial (várias seções renderizadas)


def test_api_xlsx_gera_documento_sem_excecao(client):
    resp = client.post("/api/xlsx", json=PAYLOAD_BASE)
    assert resp.status_code == 200
    assert resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_api_xlsx_aba_resumo_bate_com_payload(client):
    resp = client.post("/api/xlsx", json=PAYLOAD_BASE)
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb["Resumo"]

    linhas = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(4, 10)}
    assert linhas["Cliente"] == "123456, Cliente Teste"
    assert linhas["Assessor"] == "Denis Kinoshita"
    assert linhas["Perfil"] == "Moderada"
    assert linhas["Data de referência"] == "15/07/2026"
    assert linhas["Patrimônio"] == pytest.approx(100_000.0)
    # caixa é armazenado como fração (caixa/100) porque a célula usa number_format '0.0%'
    assert linhas["Caixa (excluído do modelo)"] == pytest.approx(0.05)


def test_api_xlsx_aba_alocacao_bate_com_desvios_enviados(client):
    resp = client.post("/api/xlsx", json=PAYLOAD_BASE)
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb["Alocação por Classe"]

    cab = [ws.cell(row=1, column=c).value for c in range(1, 7)]
    assert cab == ["Classe", "Atual (%)", "Modelo (%)", "Desvio (p.p.)", "Ação sugerida", "Valor do gap (R$)"]

    # Linha 2: Ações — real 20% vs. alvo 10%, desvio +10pp -> reduzir alocação
    row2 = [ws.cell(row=2, column=c).value for c in range(1, 7)]
    assert row2[0] == "Ações"
    assert row2[1] == pytest.approx(0.20)
    assert row2[2] == pytest.approx(0.10)
    assert row2[3] == pytest.approx(0.10)
    assert row2[4] == "Reduzir alocação"
    assert row2[5] == pytest.approx(100_000.0 * 0.10)  # patrimônio * |desvio|

    # Linha 3: Pós Fixado — real 30% vs. alvo 44%, desvio -14pp -> aumentar alocação
    row3 = [ws.cell(row=3, column=c).value for c in range(1, 7)]
    assert row3[0] == "Pós Fixado"
    assert row3[1] == pytest.approx(0.30)
    assert row3[2] == pytest.approx(0.44)
    assert row3[3] == pytest.approx(-0.14)
    assert row3[4] == "Aumentar alocação"
    assert row3[5] == pytest.approx(100_000.0 * 0.14)
